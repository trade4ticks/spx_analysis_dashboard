"""Schwab account statement -> round trips. Pure parsing, no database.

WHAT A ROUND TRIP IS HERE: a position in one symbol going 0 -> non-zero -> 0.
Not a fill, not a pairing of one buy with one sell. The strategy holds one
position at a time and exits it in whatever number of prints the book gives,
so a trip is a run of executions between two flat points and its size is the
largest position reached inside that run.

THE FAILURE THIS MODULE EXISTS TO PREVENT is an unclosed position. A day that
ends holding stock has one trip whose exit never happened, and if that trip is
silently completed at the last price -- or silently dropped -- every statistic
computed from the day is wrong in a way that reads as merely surprising. That
happened once and was not found for hours. So: an open position at the end of
the file is REPORTED and EXCLUDED, never guessed at.

SIGN CONVENTIONS. `qty` is signed as the statement writes it: BOT is positive,
SOLD is negative. Position is the running sum. A trip is long if the first leg
opened it upward, short if downward, and P&L is proceeds minus cost either way
-- which comes out right for both without a branch, because a short's opening
leg is already negative.

FEES. Schwab puts Misc Fees on the sell side only. They are attributed to the
trip containing the leg that carries them rather than spread across the day,
so a trip's own P&L nets its own costs.

NOTHING HERE TOUCHES THE DATABASE OR THE FILESYSTEM, so it can be tested
against a fabricated statement without either.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date as date_type, datetime, time as time_type


class ParseError(ValueError):
    """The file is not a statement this can read. Names what it looked for."""


# ── column resolution ───────────────────────────────────────────────────────
#
# Resolved rather than assumed, for the same reason the pipeline's schema.py
# resolves the vendor's columns: Schwab's exports differ between the web
# download and the desktop one, and a hardcoded header list fails as a
# KeyError three functions deep rather than as a sentence.
CAND_DATE = ("date", "trade date", "activity date", "settle date")
CAND_TIME = ("time", "exec time", "execution time")
CAND_TYPE = ("type", "action", "transaction type")
CAND_DESC = ("description", "desc", "activity")
CAND_REF = ("ref #", "ref#", "ref", "reference", "order #")
CAND_FEES = ("misc fees", "fees", "fee", "misc fee", "reg fee")

# Only executions. A statement carries journal entries, interest, dividends and
# balance-forward rows, and every one of them would parse to a nonsense trip if
# the description happened to contain a ticker.
TRADE_TYPES = {"trd", "trade"}

# BOT +50 FDX @334.47   /   SOLD -50 FDX @334.55
# The sign in the quantity is redundant with BOT/SOLD and they can disagree in
# a malformed row, so BOT/SOLD wins and the sign is ignored.
_EXEC_RE = re.compile(
    r"\b(?P<side>BOT|SOLD|BOUGHT|SELL|BUY)\b\s*"
    r"(?P<qty>[+-]?[\d,]+(?:\.\d+)?)\s+"
    r"(?P<symbol>[A-Z][A-Z.\-]{0,9})\s*"
    r"@\s*(?P<price>[\d,]*\.?\d+)",
    re.IGNORECASE,
)
_BUY_WORDS = {"bot", "bought", "buy"}


@dataclass
class Leg:
    ts: datetime
    symbol: str
    qty: float                       # signed: + bought, - sold
    price: float
    fees: float = 0.0
    ref: str = ""
    row: int = 0


@dataclass
class Trip:
    trade_date: date_type
    symbol: str
    entry_ts: datetime
    exit_ts: datetime
    peak_shares: float
    entry_price: float
    capital: float
    net_pnl: float
    duration_s: float
    is_long: bool
    legs: int


@dataclass
class ParseReport:
    """What the parser did. Surfaced at upload time, never only logged."""
    rows_total: int = 0
    rows_trade: int = 0
    rows_unparsed: list = field(default_factory=list)
    trips: int = 0
    unclosed: list = field(default_factory=list)
    reversals: list = field(default_factory=list)
    dates: list = field(default_factory=list)
    symbols: list = field(default_factory=list)
    columns: dict = field(default_factory=dict)

    def as_dict(self) -> dict:
        return {
            "rows_total": self.rows_total,
            "rows_trade": self.rows_trade,
            "rows_unparsed": self.rows_unparsed[:40],
            "n_unparsed": len(self.rows_unparsed),
            "trips": self.trips,
            "unclosed": self.unclosed,
            "n_unclosed": len(self.unclosed),
            "reversals": self.reversals,
            "dates": [str(d) for d in self.dates],
            "symbols": self.symbols,
            "columns": self.columns,
        }


def _norm(s) -> str:
    return str(s or "").strip().lower()


def _find(header: list[str], candidates, purpose: str, required=True):
    lower = {_norm(h): i for i, h in enumerate(header)}
    for c in candidates:
        if c in lower:
            return lower[c]
    if required:
        raise ParseError(
            f"could not find the {purpose} column. Looked for "
            f"{list(candidates)}; the file has {[str(h) for h in header]}.")
    return None


def _num(v) -> float:
    if v is None or v == "":
        return 0.0
    s = str(v).strip().replace(",", "").replace("$", "")
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    try:
        x = float(s)
    except ValueError:
        return 0.0
    return -x if neg else x


def _parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date_type):
        return v
    s = str(v).strip()
    # Schwab writes "as of" dates for corrections; the first token is the one
    # that identifies the session.
    s = s.split(" as of ")[0].strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_time(v):
    if isinstance(v, datetime):
        return v.time()
    if isinstance(v, time_type):
        return v
    s = str(v).strip().upper().replace(".", "")
    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


# ── reading the file ────────────────────────────────────────────────────────

def read_rows(data: bytes, filename: str = "") -> list[list]:
    """A statement's cells, as a list of rows. CSV or XLSX."""
    if filename.lower().endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ParseError(
                "this is an .xlsx file and openpyxl is not installed. "
                "`pip install openpyxl`, or export the statement as CSV."
            ) from exc
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(values_only=True)]

    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ParseError("could not decode the file as text; is it an .xlsx?")
    return [r for r in csv.reader(io.StringIO(text))]


def find_header(rows: list[list]) -> int:
    """The header row's index.

    Schwab prefixes the table with a title line and sometimes a blank, so the
    first row is not the header. The header is the first row that carries both
    a date-ish and a description-ish column name -- searched rather than
    assumed, because guessing row 0 fails silently by treating the title as
    column names and then finding no trades at all.
    """
    for i, row in enumerate(rows[:25]):
        names = {_norm(c) for c in row}
        if names & set(CAND_DATE) and names & set(CAND_DESC):
            return i
    raise ParseError(
        "no header row found in the first 25 lines. A Schwab statement has a "
        "row carrying both a date column and a description column.")


# ── legs -> trips ───────────────────────────────────────────────────────────

def parse_legs(rows: list[list]) -> tuple[list[Leg], ParseReport]:
    rep = ParseReport()
    if not rows:
        raise ParseError("the file is empty.")

    hi = find_header(rows)
    header = rows[hi]
    i_date = _find(header, CAND_DATE, "date")
    i_desc = _find(header, CAND_DESC, "description")
    i_type = _find(header, CAND_TYPE, "type", required=False)
    i_time = _find(header, CAND_TIME, "time", required=False)
    i_ref = _find(header, CAND_REF, "ref", required=False)
    i_fees = _find(header, CAND_FEES, "fees", required=False)
    rep.columns = {"date": i_date, "time": i_time, "type": i_type,
                   "description": i_desc, "ref": i_ref, "fees": i_fees}

    legs: list[Leg] = []
    for n, row in enumerate(rows[hi + 1:], start=hi + 2):
        if not row or all(c in (None, "") for c in row):
            continue
        rep.rows_total += 1

        if i_type is not None:
            if _norm(row[i_type] if i_type < len(row) else "") not in TRADE_TYPES:
                continue
        rep.rows_trade += 1

        desc = str(row[i_desc]) if i_desc < len(row) else ""
        m = _EXEC_RE.search(desc)
        d = _parse_date(row[i_date]) if i_date < len(row) else None
        if not m or d is None:
            # Kept with its line number and its text. "12 rows did not parse"
            # is unactionable; the row itself is what says whether the parser
            # is wrong or the statement carries something new.
            rep.rows_unparsed.append({"line": n, "text": desc[:120]})
            continue

        t = _parse_time(row[i_time]) if i_time is not None and i_time < len(row) else None
        ts = datetime.combine(d, t) if t else datetime.combine(d, time_type(0, 0))
        qty = abs(_num(m.group("qty")))
        if m.group("side").lower() in _BUY_WORDS:
            signed = qty
        else:
            signed = -qty
        legs.append(Leg(
            ts=ts, symbol=m.group("symbol").upper(), qty=signed,
            price=_num(m.group("price")),
            fees=abs(_num(row[i_fees])) if i_fees is not None and i_fees < len(row) else 0.0,
            ref=str(row[i_ref]) if i_ref is not None and i_ref < len(row) else "",
            row=n,
        ))

    if not legs and rep.rows_trade:
        raise ParseError(
            f"{rep.rows_trade} rows had a trade type but none matched an "
            f"execution like 'BOT +50 FDX @334.47'.")
    return legs, rep


def build_trips(legs: list[Leg], rep: ParseReport) -> list[Trip]:
    """Round trips, one per 0 -> non-zero -> 0 excursion per symbol per day.

    Grouped by (date, symbol) and walked in time order. A leg that would carry
    the position THROUGH zero is split at zero: the part that closes the
    existing trip belongs to it, the remainder opens the next one. That case
    should not arise in a one-position-at-a-time strategy, so it is also
    recorded -- an unexpected reversal is worth seeing rather than smoothing.
    """
    trips: list[Trip] = []
    by_key: dict = {}
    for lg in legs:
        by_key.setdefault((lg.ts.date(), lg.symbol), []).append(lg)

    for (d, sym), group in sorted(by_key.items()):
        group.sort(key=lambda x: (x.ts, x.row))
        pos = 0.0
        open_legs: list[tuple[datetime, float, float, float]] = []   # ts, qty, price, fees

        def close(open_legs):
            """One completed excursion -> a Trip."""
            buys = sum(q for _, q, _, _ in open_legs if q > 0)
            sells = -sum(q for _, q, _, _ in open_legs if q < 0)
            cost = sum(q * p for _, q, p, _ in open_legs if q > 0)
            proceeds = -sum(q * p for _, q, p, _ in open_legs if q < 0)
            fees = sum(f for _, _, _, f in open_legs)
            # Running position through the excursion, for the largest size the
            # trip ever carried -- which is the capital at risk, not the sum
            # of the legs.
            run, peak = 0.0, 0.0
            for _, q, _, _ in open_legs:
                run += q
                peak = max(peak, abs(run))
            first_qty = open_legs[0][1]
            is_long = first_qty > 0
            opening = [(q, p) for _, q, p, _ in open_legs
                       if (q > 0) == is_long]
            oq = sum(q for q, _ in opening) or 1.0
            entry_price = sum(q * p for q, p in opening) / oq
            ts0, ts1 = open_legs[0][0], open_legs[-1][0]
            return Trip(
                trade_date=d, symbol=sym, entry_ts=ts0, exit_ts=ts1,
                peak_shares=peak, entry_price=abs(entry_price),
                capital=peak * abs(entry_price),
                # proceeds - cost works for both directions without a branch:
                # a short's opening leg is already the negative one.
                net_pnl=proceeds - cost - fees,
                duration_s=(ts1 - ts0).total_seconds(),
                is_long=is_long, legs=len(open_legs),
            )

        for lg in group:
            q = lg.qty
            if pos != 0 and (pos > 0) != (q > 0) and abs(q) > abs(pos):
                # Through zero. Split so the closing part belongs to the trip
                # it closes; the remainder opens the next.
                closing = -pos
                remainder = q - closing
                open_legs.append((lg.ts, closing, lg.price, lg.fees))
                trips.append(close(open_legs))
                rep.reversals.append(
                    {"date": str(d), "symbol": sym, "line": lg.row,
                     "note": "a single execution carried the position through "
                             "zero and was split"})
                open_legs = [(lg.ts, remainder, lg.price, 0.0)]
                pos = remainder
                continue

            open_legs.append((lg.ts, q, lg.price, lg.fees))
            pos += q
            if abs(pos) < 1e-9:
                trips.append(close(open_legs))
                open_legs = []
                pos = 0.0

        if open_legs:
            # NEVER completed at the last price and never silently dropped.
            # An unclosed position is a fact about the day, and inventing an
            # exit is what corrupted a session's statistics before.
            rep.unclosed.append({
                "date": str(d), "symbol": sym,
                "shares": round(pos, 4),
                "since": open_legs[0][0].isoformat(timespec="seconds"),
                "legs": len(open_legs),
            })

    rep.trips = len(trips)
    rep.dates = sorted({t.trade_date for t in trips})
    rep.symbols = sorted({t.symbol for t in trips})
    return trips


def parse_statement(data: bytes, filename: str = "") -> tuple[list[Trip], ParseReport]:
    legs, rep = parse_legs(read_rows(data, filename))
    return build_trips(legs, rep), rep


# ── daily aggregate ─────────────────────────────────────────────────────────

def daily_rows(trips: list[Trip]) -> list[dict]:
    """One row per (date, symbol).

    ATTENTION MINUTES IS WALL CLOCK, first entry to last exit, NOT the sum of
    hold durations. That is the number the strategy is actually rationed by:
    a name that took 96 minutes of the session for 14 round trips consumed 96
    minutes whether or not a position was open in each of them, and $/min
    against summed hold time would rate it as though the waiting were free.
    """
    by: dict = {}
    for t in trips:
        by.setdefault((t.trade_date, t.symbol), []).append(t)

    out = []
    for (d, sym), ts in sorted(by.items()):
        pnl = sum(x.net_pnl for x in ts)
        wins = sum(1 for x in ts if x.net_pnl > 0)
        first = min(x.entry_ts for x in ts)
        last = max(x.exit_ts for x in ts)
        minutes = (last - first).total_seconds() / 60.0
        out.append({
            "trade_date": d, "symbol": sym,
            "trips": len(ts),
            "net_pnl": pnl,
            "win_rate": wins / len(ts) if ts else None,
            "attention_minutes": minutes,
            # None, not zero, when every trip started and ended in the same
            # second. Zero would be a rate and this is an absence of one.
            "dollars_per_min": (pnl / minutes) if minutes > 0 else None,
            "trips_per_min": (len(ts) / minutes) if minutes > 0 else None,
            "median_hold_s": sorted(x.duration_s for x in ts)[len(ts) // 2],
            "shares": sum(x.peak_shares for x in ts),
        })
    return out
